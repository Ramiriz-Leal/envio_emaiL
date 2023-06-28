import pyodbc
import smtplib
import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

# Configurações de conexão com o banco de dados
server = 'IP BANCO DE DADOS'
database = 'BANCO DE DADOS'
username = 'USUARIO'
password = 'SENHA'
consulta_sql = '''
COMANDO SELECT SQL SERVER
'''

# Configurações de conexão com o servidor de e-mail
smtp_host = 'smtp.dominio.com.br'
smtp_port = 'porta'
smtp_username = 'email@email.com.br'
smtp_password = 'senha'

# Configurações dos e-mails
remetente = 'email@email.com.br'
destinatarios = ['email@email.com.br','email@email.com.br']
assunto = 'Assunto do email'

# Conectando-se ao banco de dados e executando a consulta
conexao = pyodbc.connect(f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}')
cursor = conexao.cursor()
cursor.execute(consulta_sql)
resultados = cursor.fetchall()
conexao.close()

# Converting results of the query to a list of tuples
resultados = [tuple(row) for row in resultados]

# Obter a data e hora atual
data_hora_atual = datetime.datetime.now()

# Converter para uma string formatada
data_hora_formatada = data_hora_atual.strftime('%d/%m/%Y %H:%M:%S')

# Convertendo resultados da consulta para um DataFrame pandas
columns = [column[0] for column in cursor.description]
df = pd.DataFrame(resultados, columns=columns)

# Criando um novo arquivo Excel na mesma pasta python com o DataFrame
wb = Workbook()
ws = wb.active

# Escrevendo os nomes das colunas no arquivo Excel e aplicando os tamanhos
for col_num, column_title in enumerate(columns, 1):
    col_letter = get_column_letter(col_num)
    cell = ws[f'{col_letter}1']
    cell.value = column_title
    cell.font = cell.font.copy(bold=True, size=12)
    cell.alignment = cell.alignment.copy(horizontal='center')

# Escrevendo os dados no arquivo Excel
for row_data in resultados:
    ws.append(row_data)

# Ajustando as larguras das colunas do resultados com base no tamanho dos dados
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    adjusted_width = (length + 1) * 1.0
    ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

# Salvando o arquivo Excel
excel_file = BytesIO()
wb.save(excel_file)
excel_file.seek(0)

# Preparando o conteúdo do e-mail
mensagem = MIMEMultipart()
mensagem['From'] = remetente
mensagem['To'] = ', '.join(destinatarios)
mensagem['Subject'] = assunto

# Adicionando o corpo do e-mail à mensagem
corpo_email = f'<h2>Resultado de acordo com a data: {data_hora_formatada}</h2>'
corpo_email += '<br>'
corpo_email += '<br>======================= E-mail automático ======================='
mensagem.attach(MIMEText(corpo_email, 'html'))

# Adicionando o arquivo Excel como anexo
anexo = MIMEBase('application', 'octet-stream')
anexo.set_payload(excel_file.read())
encoders.encode_base64(anexo)
anexo.add_header('Content-Disposition', 'attachment', filename='usuarios_senior_erp.xlsx')
mensagem.attach(anexo)

# Enviando o e-mail
with smtplib.SMTP_SSL(smtp_host, smtp_port) as smtp:
    smtp.login(smtp_username, smtp_password)
    smtp.send_message(mensagem)